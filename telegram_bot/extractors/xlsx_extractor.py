"""
XLSX/CSV text extraction: convert sheets to readable text.
"""
from __future__ import annotations

import io
import csv
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None


def extract_text_from_xlsx(data: bytes) -> str:
    """Extract text from .xlsx: all sheets as tab-separated text."""
    if openpyxl is not None:
        return _extract_with_openpyxl(data)
    if pd is not None:
        return _extract_with_pandas(data)
    raise RuntimeError("Install openpyxl or pandas: pip install openpyxl pandas")


def _extract_with_openpyxl(data: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"--- {sheet_name} ---")
            for row in ws.iter_rows(values_only=True):
                line = "\t".join(str(c) if c is not None else "" for c in row)
                if line.strip():
                    parts.append(line)
            parts.append("")
    finally:
        wb.close()
    return "\n".join(parts).strip()


def _extract_with_pandas(data: bytes) -> str:
    if pd is None:
        return ""
    xl = pd.ExcelFile(io.BytesIO(data))
    parts: list[str] = []
    for name in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=name, header=None)
        parts.append(f"--- {name} ---")
        parts.append(df.to_string(index=False, header=False))
        parts.append("")
    return "\n".join(parts).strip()


def extract_text_from_csv(data: bytes, encoding: str = "utf-8") -> str:
    """Extract text from CSV; try utf-8 then cp1251."""
    for enc in (encoding, "utf-8", "cp1251", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")

    reader = csv.reader(io.StringIO(text))
    lines = []
    for row in reader:
        lines.append("\t".join(row))
    return "\n".join(lines)
